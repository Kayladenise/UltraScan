import sys
import os
from tkinter import *
from PIL import Image, ImageTk
from openpyxl.utils.dataframe import dataframe_to_rows
from ttkbootstrap.dialogs.dialogs import Querybox
from ttkbootstrap.toast import ToastNotification
import ttkbootstrap as ttkb
import tkinter.filedialog as filedialog
import numpy as np
import pandas as pd
import cv2
import openpyxl
import datetime
import time
import segment as sg

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

root = Tk()
root.title('UltraScan')
root.geometry("1600x900")
root.resizable(False, False)
style = ttkb.Style(theme="litera")
style.configure("info.Treeview", rowheight=40)
# style.configure("Treeview", font=("Helvetica",))
style.configure("Treeview.Heading", font=("Helvetica", 22))

icon_path = resource_path("images/main_icon.png")
if os.path.exists(icon_path):
    main_icon = ImageTk.PhotoImage(file=icon_path)
    root.iconphoto(False, main_icon)
else:
    print(f"Error: The file {icon_path} does not exist.")

headerFont = "Helvetica"
bodyFont = "Verdana"

btn_style = ttkb.Style().configure(style="my.TButton", font=(bodyFont, 30))
btn_style1 = ttkb.Style().configure(style="primary.TButton", font=(headerFont, 20))
btn_style2 = ttkb.Style().configure(style="light.TButton", font=(headerFont, 20))

canvas = Canvas(root, width=1600, height=900, highlightthickness=0)
canvas.pack(fill="both", expand=True)


def startScreen():
    print("Starting screen")
    canvas.delete("all")
    drawBackground("LANDING_PAGE.png")
    
    # get started button
    start_img = Image.open(resource_path("images/BUTTON_GET_STARTED.png"))
    start_img = start_img.resize((415, 69))
    start_img_tk = ImageTk.PhotoImage(start_img)
    
    start_btn = canvas.create_image(1035, 655, image=start_img_tk, anchor="center")
    canvas.tag_bind(start_btn, "<Button-1>", lambda event: segmentScreen())
    print("Start screen setup complete")

def segmentScreen():
    print("Segment screen")
    canvas.delete("all")
    sideBar("segment")
    drawBackground("SEGMENT_PAGE.png")

    select_img = Image.open(resource_path("images/BUTTON_SELECT_IMAGE.png"))
    select_img = select_img.resize((415, 69))
    select_img_tk = ImageTk.PhotoImage(select_img)

    selectImg_btn = canvas.create_image(700, 750, image=select_img_tk, anchor="center")
    canvas.tag_bind(selectImg_btn, "<Button-1>", lambda event: selectImg())
    canvas.select_image = select_img_tk  # reference to avoid garbage collection

    question_img = Image.open(resource_path("images/BUTTON_QUESTION.png"))
    question_img = question_img.resize((40, 38))
    question_img_tk = ImageTk.PhotoImage(question_img)
    questionImg_btn = canvas.create_image(1400, 100, image=question_img_tk, anchor="center")
    canvas.tag_bind(questionImg_btn, "<Button-1>", lambda event: userManualScreen())
    canvas.question_image = question_img_tk 

    print("Home screen setup complete")

def recordScreen():
    def viewImg():
        selection = my_tree.selection()
        if selection:
            item = my_tree.item(selection[0])
            name, pixelsize, angle, bpd, ofd, hc, age, date = item['values']
            viewScreen(name, pixelsize, angle, bpd, ofd, hc, age, date)
        else:
            notifToast("View Image Failed",
                "Please select a row first.",
                "danger")
    def delRecord():
        selection = my_tree.selection()
        if selection:
            item = my_tree.item(selection[0])
            name, pixelsize, _, _, _, _, _, date  = item['values']
            deleteRecord(name, pixelsize, date)
        else:
            notifToast("Deleting Failed",
                "Please select a row first.",
                "danger")

    print("Show Record screen")
    canvas.delete("all")
    sideBar("records")
    # check if excel exists and if empty
    file_path = 'Fetal Head.xlsx'
    if not os.path.exists(file_path):
        drawBackground("RECORD_PAGE_EMPTY.png")
        return
    else:
        df = pd.read_excel(file_path)
        if df.shape[0] <= 0:
            drawBackground("RECORD_PAGE_EMPTY.png")
            return
    drawBackground("RECORD_PAGE.png")

    records = df.to_dict('records')

    columns = ("name", "pixelsize", "angle", "bpd", "ofd", "hc", "age", "date")
    my_tree = ttkb.Treeview(root, bootstyle="info", style="info.Treeview", columns=columns, show="headings", selectmode="browse")

    for column in columns:
        if column in ["bpd", "ofd", "hc", "angle"]:
            my_tree.column(column, width=100)
        else:
            my_tree.column(column, width=200)
    
    # disable column resizing
    my_tree.bind('<Motion>', 'break')

    canvas.create_window(300, 170, anchor="nw", window=my_tree)
    
    my_tree.heading('name', text="Name", anchor="w")
    my_tree.heading('pixelsize', text="Pixel Size", anchor="w")
    my_tree.heading('angle', text="Angle", anchor="w")
    my_tree.heading('bpd', text="BPD", anchor="w")
    my_tree.heading('ofd', text="OFD", anchor="w")
    my_tree.heading('hc', text="HC", anchor="w")
    my_tree.heading('age', text="Age (wk)", anchor="w")
    my_tree.heading('date', text="Date Added", anchor="w")

    #add data to treeview
    for record in records:
        my_tree.insert("", END, values=list(record.values()))

    # Load button images
    view_img = Image.open(resource_path("images/BUTTON_VIEW_IMAGE.png"))
    view_img = view_img.resize((219, 69))
    view_img_tk = ImageTk.PhotoImage(view_img)

    delete_img = Image.open(resource_path("images/BUTTON_DELETE_RECORD.png"))
    delete_img = delete_img.resize((219, 69))
    delete_img_tk = ImageTk.PhotoImage(delete_img)

    # Create buttons with images
    view_btn = canvas.create_image(750, 720, image=view_img_tk, anchor="center")
    canvas.tag_bind(view_btn, "<Button-1>", lambda event:viewImg())
    canvas.calculate_image = view_img_tk

    delete_btn = canvas.create_image(1000, 720, image=delete_img_tk, anchor="center")
    canvas.tag_bind(delete_btn, "<Button-1>", lambda event:delRecord())
    canvas.delete_img = delete_img_tk


def userManualScreen():
    print("User Manual screen")
    canvas.delete("all")
    drawBackground("MANUAL_PAGE_EMPTY.png")
    
    proceed_img = Image.open(resource_path("images/BUTTON_UNDERSTAND.png"))
    proceed_img = proceed_img.resize((415, 69))
    proceed_img_tk = ImageTk.PhotoImage(proceed_img)
    proceed_btn = canvas.create_image(1035, 750, image=proceed_img_tk, anchor="center")
    canvas.tag_bind(proceed_btn, "<Button-1>", lambda event: segmentScreen())
    canvas.proceed_image = proceed_img_tk

def resultScreen(image):
    print("Result screen")
    
    def calculateParams(a, b, angle, image):
        try:
            nonlocal bpd, ofd, hc, gestAge
            pixelsize = float(my_entry.get())
            if pixelsize <= 0:
                raise ValueError
            bpd, ofd, hc, gestAge = getParameters(pixelsize, a, b)
            print(f"Parameters calculated: bpd={bpd}, ofd={ofd}, hc={hc}, gestAge={gestAge}, angle={angle}")
            canvas.itemconfig(angle_item, text=f'{np.round(angle, 2)} degrees')
            bpd = np.round(bpd, 2)
            ofd = np.round(ofd, 2)
            hc = np.round(hc, 2)

            canvas.itemconfig(bpd_item, text=f'{bpd} mm')
            canvas.itemconfig(ofc_item, text=f'{ofd} mm')
            canvas.itemconfig(hc_item, text=f'{hc} cm')
            if gestAge in ["<8 Menstrual wk","Abnormal"]:
                canvas.itemconfig(gestAge_item, text=f'{gestAge}')
            else:
                canvas.itemconfig(gestAge_item, text=f'{gestAge} weeks')

            save_btn = canvas.create_image(680, 690, image=save_img_tk, anchor="center")
            canvas.tag_bind(save_btn, "<Button-1>", lambda event: saveInfo(pixelsize, angle, bpd, ofd, hc, gestAge, save_btn, image))
        except ValueError:
            notifToast("Pixel Size Warning",
                    "Please input an integer greater than 0.",
                    "danger")
            print("ValueError in calculateParams")


    bpd, ofd, hc, gestAge, angle = None, None, None, None, None

    # Segmented image
    start_time = time.process_time()
    result = sg.segment(image)
    end_time = time.process_time()
    elapsed_time = end_time - start_time
    print(f"Process took {elapsed_time:.4f} seconds to complete")
    if result is None:
        print("An error occurred during segmentation")
        notifToast("Segmentation Failed",
            "Please choose a suitable image.",
            "danger")
        return
    else:
        segmented_image, semi_axes_a, semi_axes_b, angle = result

    # Background image
    canvas.delete("all")
    sideBar("segment")
    drawBackground("RESULT_PAGE.png")

    print(f"Segmented image: semi_axes_a={semi_axes_a}, semi_axes_b={semi_axes_b}, angle={angle}")
    segmented_image_pil = Image.fromarray(segmented_image)  # Convert image to PIL format
    segmented_image = segmented_image_pil.resize((640, 400))  # Resize image
    segmented_image_tk = ImageTk.PhotoImage(segmented_image)  # Convert image to PhotoImage
    canvas.create_image(1240, 400, image=segmented_image_tk, anchor="center")  # Adjusted position to the right
    canvas.segmented_image = segmented_image_tk  # Keep a reference to avoid garbage collection

    notifToast("Segmentation Successful!",
            "The image has been successfully segmented.",
            "success")

    # Load button images
    calculate_img = Image.open(resource_path("images/BUTTON_CALCULATE.png"))
    calculate_img = calculate_img.resize((219, 69))
    calculate_img_tk = ImageTk.PhotoImage(calculate_img)

    save_img = Image.open(resource_path("images/BUTTON_SAVE.png"))
    save_img = save_img.resize((219, 69))
    save_img_tk = ImageTk.PhotoImage(save_img)

    # Create buttons with images
    calculate_btn = canvas.create_image(420, 690, image=calculate_img_tk, anchor="center")
    canvas.tag_bind(calculate_btn, "<Button-1>", lambda event: calculateParams(semi_axes_a, semi_axes_b, angle, segmented_image_pil))
    canvas.calculate_image = calculate_img_tk


    my_entry = ttkb.Entry(canvas,
                          bootstyle="info",
                          font=(bodyFont, 18),
                          width=14)

    canvas.create_window(470, 200, anchor="nw", window=my_entry)

    canvas.create_text(440, 200, text="Pixel Size:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(385, 275, text="Angle:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(385, 325, text="BPD:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(385, 375, text="OFD:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(385, 425, text="HC:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(290, 500, text="Gestational Age:", anchor="nw", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(1100, 690, text=f"Time elapsed: {elapsed_time:.4f}s", anchor="nw", font=(bodyFont, 20, "bold"), fill="black")

    angle_item = canvas.create_text(440, 275, anchor="nw", text="", font=(bodyFont, 20), fill="black")
    bpd_item = canvas.create_text(440, 325, anchor="nw", text="", font=(bodyFont, 20), fill="black")
    ofc_item = canvas.create_text(440, 375, anchor="nw", text="", font=(bodyFont, 20), fill="black")
    hc_item = canvas.create_text(440, 425, anchor="nw", text="", font=(bodyFont, 20), fill="black")
    gestAge_item = canvas.create_text(440, 550, text="", anchor="nw", font=(bodyFont, 20), fill="black")

    print("Segment screen setup complete")

def viewScreen(file, pixelsize, angle, bpd, ofd, hc, age, date):
    print("View screen")
    # Background image
    canvas.delete("all")
    sideBar("records")
    drawBackground("VIEW_PAGE.png")

    img = Image.open(f"segmented/{file}.png")
    img = img.resize((640, 400))
    img_tk = ImageTk.PhotoImage(img)
    canvas.create_image(1240, 400, image=img_tk, anchor="center")
    canvas.image = img_tk
    
    # Load button images
    delete_img = Image.open(resource_path("images/BUTTON_DELETE_RECORD.png"))
    delete_img = delete_img.resize((219, 69))
    delete_img_tk = ImageTk.PhotoImage(delete_img)

    # Create buttons with images
    delete_btn = canvas.create_image(420, 690, image=delete_img_tk, anchor="center")
    canvas.tag_bind(delete_btn, "<Button-1>", lambda event:deleteRecord(file, pixelsize, date))
    canvas.calculate_image = delete_img_tk

    canvas.create_text(440, 200, text="Pixel Size:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(385, 275, text="Angle:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(385, 325, text="BPD:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(385, 375, text="OFD:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(385, 425, text="HC:", anchor="ne", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(290, 500, text="Gestational Age:", anchor="nw", font=(bodyFont, 20, "bold"), fill="black")
    canvas.create_text(1100, 620, text=f"segmented/{file}.png", anchor="nw", font=(bodyFont, 20, "bold"), fill="black")

    canvas.create_text(460, 200, anchor="nw", text=pixelsize, font=(bodyFont, 20), fill="black")
    canvas.create_text(440, 275, anchor="nw", text=f"{angle} degrees", font=(bodyFont, 20), fill="black")
    canvas.create_text(440, 325, anchor="nw", text=f"{bpd} mm", font=(bodyFont, 20), fill="black")
    canvas.create_text(440, 375, anchor="nw", text=f"{ofd} mm", font=(bodyFont, 20), fill="black")
    canvas.create_text(440, 425, anchor="nw", text=f"{hc} cm", font=(bodyFont, 20), fill="black")
    
    if age not in ["<8 Menstrual wk","Abnormal"]:
        age = f"{age} weeks"
    canvas.create_text(440, 550, anchor="nw", text=age, font=(bodyFont, 20), fill="black")
    


def getParameters(pixelSize, a, b):
    print("Calculating parameters")
    semi_axes_a_mm = a * pixelSize / 2
    semi_axes_b_mm = b * pixelSize / 2

    bpd = semi_axes_a_mm * 2
    ofd = semi_axes_b_mm * 2

    hc = 1.62 * (bpd + ofd)
    hc = np.round(hc / 10, 3)

    conditions = [
        (hc < 8.00),
        (hc >= 8.00) & (hc <= 9.00),  # week 13
        (hc > 9.01) & (hc <= 10.49),  # week 14
        (hc > 10.50) & (hc <= 12.49),  # week 15
        (hc > 12.50) & (hc <= 13.49),  # week 16
        (hc > 13.50) & (hc <= 14.99),  # week 17
        (hc > 15.00) & (hc <= 16.49),  # week 18
        (hc > 16.50) & (hc <= 17.49),  # week 19
        (hc > 17.50) & (hc <= 18.99),  # week 20
        (hc > 19.00) & (hc <= 19.99),  # week 21
        (hc > 20.00) & (hc <= 20.99),  # week 22
        (hc > 21.00) & (hc <= 22.49),  # week 23
        (hc > 22.50) & (hc <= 22.99),  # week 24
        (hc >= 23.00) & (hc <= 23.99),  # week 25
        (hc > 24.00) & (hc <= 24.79),  # week 26
        (hc > 24.80) & (hc <= 25.60),  # week 27
        (hc > 25.61) & (hc <= 26.75),  # week 28
        (hc > 26.76) & (hc <= 27.75),  # week 29
        (hc > 27.76) & (hc <= 28.85),  # week 30
        (hc > 28.86) & (hc <= 29.60),  # week 31
        (hc > 29.61) & (hc <= 30.40),  # week 32
        (hc > 30.41) & (hc <= 31.20),  # week 33
        (hc > 31.21) & (hc <= 31.80),  # week 34
        (hc > 31.81) & (hc <= 32.50),  # week 35
        (hc > 32.51) & (hc <= 33.00),  # week 36
        (hc > 33.01) & (hc <= 33.70),  # week 37
        (hc > 33.71) & (hc <= 34.20),  # week 38
        (hc > 34.21) & (hc <= 35.00),  # week 39
        (hc > 35.00) & (hc <= 36.00),  # week 40
        (hc > 36)
    ]

    values = ['<8 Menstrual wk', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22',
              '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39',
              '40', 'Abnormal']

    return bpd, ofd, hc, np.select(conditions, values)

def selectImg():
    file_path = filedialog.askopenfilename(title="Select Image",
                                           filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])
    if file_path:
        print("Selected file:", file_path)  # to be edited
        imgNp = cv2.imread(file_path, cv2.IMREAD_COLOR)

        img = Image.open(file_path)
        img = img.resize((640, 360))
        img_tk = ImageTk.PhotoImage(img)

        canvas.create_image(910, 470, image=img_tk, anchor="center")
        canvas.image = img_tk

        canvas.create_text(910, 675, text="Image successfully chosen!", anchor="center", font=(bodyFont, 15), fill="black")

        segment_img = Image.open(resource_path("images/BUTTON_SEGMENT.png"))
        segment_img = segment_img.resize((415, 69))
        segment_img_tk = ImageTk.PhotoImage(segment_img)

        segment_btn = canvas.create_image(1150, 750, image=segment_img_tk, anchor="center")
        canvas.tag_bind(segment_btn, "<Button-1>", lambda event: resultScreen(imgNp))
        canvas.segment_image = segment_img_tk

def saveInfo(pixelsize, angle, bpd, ofd, hc, gestAge, button, image):
    while True:
        name = Querybox.get_string(prompt='Enter name to be saved:', title='Save Info')
        if name is None:
            return
        elif name == '':
            notifToast("Save Unsuccessful",
                "Name should not be blank.",
                "danger")
        elif len(name)>16:
            notifToast("Save Unsuccessful",
                "Name should not be greater than 16 characters.",
                "danger")
        else:
            break

    date = datetime.date.today().strftime("%m/%d/%Y")  # Convert date to string
    data = {'Name': [name], 'Pixel Size':[pixelsize], 'Angle': [round(angle, 2)], 'BPD': [bpd], 'OFD': [ofd], 'HC': [hc],
            'Age (wk)': [np.array2string(gestAge).strip("'")], 'Date Added': [date]}
    df = pd.DataFrame(data)
    file_path = "Fetal Head.xlsx"
    # Check if file exists
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
    else:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # Save the workbook
    wb.save(file_path)
    notifToast("Save Successful",
               "Successfully saved at Fetal Head.xlsx",
               "success")
    canvas.itemconfig(button, state=HIDDEN)

    image_path = f'segmented/{name}.png'
    # Create the directory if it doesn't exist
    directory = os.path.dirname(image_path)
    if not os.path.exists(directory):
        os.makedirs(directory)
    image.save(image_path, 'PNG')

def sideBar(screen):
    btn1 = ttkb.Button(text="Segment", width=13, padding=9, style="primary.TButton", command=segmentScreen,)
    btn2 = ttkb.Button(text="Records", width=13, padding=9, style="primary.TButton", command=recordScreen)
    
    
    if screen == "segment":
        btn1.configure(style="light.TButton")
    if screen == "records":
        btn2.configure(style="light.TButton")
    canvas.create_window(0, 110, anchor = "nw", window=btn1)
    canvas.create_window(0, 165, anchor = "nw", window=btn2)

def notifToast(title, msg, style):
    ToastNotification(title=title,
                                message=msg,
                                duration=3000,
                                bootstyle=style).show_toast()

def drawBackground(file):
    global bg_image_tk
    bg_image = Image.open(resource_path(f"images/{file}"))
    bg_image = bg_image.resize((1600, 900))
    bg_image_tk = ImageTk.PhotoImage(bg_image)
    canvas.create_image(800, 450, image=bg_image_tk, anchor="center")
    canvas.image = bg_image_tk

def deleteRecord(name, pixelsize, date):
    found = False
    wb = openpyxl.load_workbook('Fetal Head.xlsx')
    sheet = wb.active
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if str(name) in row and float(pixelsize) in row and date in row:
            os.remove(f'segmented/{name}.png')
            sheet.delete_rows(row_idx)
            found = True
            notifToast("Record Deleted",
                f"Successfully deleted {name}",
                "success")
            break
    if not found:
        notifToast("Deleting Failed",
            "Data not found in Fetal Head.xlsx",
            "danger")
    wb.save('Fetal Head.xlsx')
    recordScreen()

startScreen()


root.mainloop()
